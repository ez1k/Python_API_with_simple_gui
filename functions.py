import pandas as pd
import os
import requests
import PySimpleGUI as sg
from Classes import My_Weather
import openpyxl
from openpyxl.chart import LineChart, Reference
import openpyxl.utils.cell as ut
import xlsxwriter as writer
def get_data(country, city_name):
    WEATHER_URL = "http://api.openweathermap.org/data/2.5/weather?q="
    api_key = "97d8990dc09a98a2a1479ec044bac0eb"
    complete_url = WEATHER_URL + city_name + "," + country + "&APPID=" + api_key + "&units=metric"

    response = requests.get(complete_url)
    json_Data = response.json()

    if json_Data["cod"] != "404" and json_Data["cod"] != "400":

        coordinates = json_Data["coord"]
        latitude = coordinates['lat']
        longitude = coordinates['lon']

        wind_data = json_Data['wind']
        wind_speed = wind_data['speed']

        main_data_weather = json_Data["main"]
        current_temperature = main_data_weather["temp"]
        current_pressure = main_data_weather["pressure"]
        current_humidity = main_data_weather["humidity"]

        weather_description = json_Data["weather"]
        weather_description = weather_description[0]["description"]

        actual_weather = My_Weather(current_temperature, wind_speed, current_pressure, current_humidity,
                                    weather_description)
    else:
        actual_weather = My_Weather("no data", "no data", "no data", "no data",
                                    "no data")
        sg.popup('Wrong input data')
        

    return actual_weather


def collect_data(df):
    for index, row in df.iterrows():
        w1 = get_data(row['country'], row['city'])
        df.loc[index, 'Temperature'] = w1.temperature
        df.loc[index, 'Wind Speed'] = w1.wind_speed
        df.loc[index, 'Pressure'] = w1.pressure
        df.loc[index, 'Humidity'] = w1.humidity
        df.loc[index, 'Description'] = w1.description
    return df
    
def create_chart(filename, chart_title, x_title, y_title, x_column, y_column, plot_place):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    

    chart = LineChart()
    chart.title = chart_title
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title

    xvalues = Reference(ws, min_col=ut.column_index_from_string(x_column), min_row=2, max_row=ws.max_row)
    yvalues = Reference(ws, min_col=ut.column_index_from_string(y_column), min_row=1, max_row=ws.max_row)

    chart.add_data(yvalues, titles_from_data=True)
    chart.set_categories(xvalues)
    ws.add_chart(chart, plot_place)


    wb.save(filename)


def create_Excel(df, local_path):
    writer = pd.ExcelWriter(local_path + "results.xlsx", engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    worksheet = writer.sheets['Sheet1']
    for column in df:
        column_length = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        worksheet.set_column(col_idx, col_idx, column_length)
    writer.save()